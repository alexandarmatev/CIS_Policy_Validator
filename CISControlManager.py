from typing import Tuple, Dict, NamedTuple, List
from openpyxl.worksheet.worksheet import Worksheet
from workbook_management.excel_workbook_manager import ExcelOpenWorkbook
from DataModels import CISControl, CISControlFamily
from collections import namedtuple
from collections import Counter
from workbook_management.interfaces import IWorkbookLoader, IConfigLoader
from utils.validation_utils import validate_and_return_file_path


class CISControlManager(ExcelOpenWorkbook):
    def __init__(self, workbook_loader: IWorkbookLoader, config_loader: IConfigLoader, workbook_path: str, config_path: str):
        self._workbook_path = validate_and_return_file_path(workbook_path, 'xlsx')
        self._config_path = validate_and_return_file_path(config_path, 'json')
        super().__init__(workbook_loader, config_loader)
        self._control_families = {}
        self._cache = {'All Controls': []}
        self._populate_controls_cache()

    def load_workbook(self):
        return self._workbook_loader.load(self._workbook_path)

    def load_config(self):
        return self._config_loader.load(self._config_path)[__class__.__name__]

    @property
    def worksheet_name(self) -> str:
        return self._config['WORKSHEET_NAME']

    @property
    def asset_type(self) -> str:
        return self._config['ASSET_TYPE']

    @property
    def domain(self) -> str:
        return self._config['DOMAIN']

    @property
    def control_family_id(self) -> str:
        return self._config['CONTROL_FAMILY_ID']

    def _get_worksheet_scope_headers(self) -> Tuple[Worksheet, Dict[str, int]]:
        worksheet_name = self._validate_and_return_sheet_name(self.worksheet_name)
        worksheet = self._workbook[worksheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {title: index for index, title in enumerate(header_row)}
        return worksheet, column_indices

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> NamedTuple:
        RowData = namedtuple('RowData', ['safeguard_id', 'asset_type', 'domain', 'title', 'description', 'control_family_id', 'is_family'])
        if self._validate_column_titles(column_indices, self.required_column_titles):
            safeguard_ids = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                safeguard_id = str(row[column_indices[self.safeguard]])
                if safeguard_id in safeguard_ids:
                    safeguard_id += '0'
                safeguard_ids.add(safeguard_id)
                asset_type = row[column_indices[self.asset_type]]
                domain = row[column_indices[self.domain]]
                title = row[column_indices[self.title]]
                description = row[column_indices[self.description]]
                control_family_id = None
                is_family = False

                if not asset_type:
                    is_family = True
                    control_family_id = str(row[column_indices[self.control_family_id]])

                yield RowData(safeguard_id, asset_type, domain, title, description, control_family_id, is_family)

    def _populate_controls_cache(self):
        worksheet, column_indices = self._get_worksheet_scope_headers()
        worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
        for row_data in worksheet_row_attrs:
            if row_data.is_family:
                control_family_id = row_data.control_family_id.strip()
                self._control_families[control_family_id] = CISControlFamily(title=row_data.title, description=row_data.description)
            else:
                self._cache['All Controls'].append(CISControl(safeguard_id=row_data.safeguard_id, asset_type=row_data.asset_type,
                                                              domain=row_data.domain, title=row_data.title, description=row_data.description))

    def get_all_controls(self) -> List[CISControl]:
        return self._cache['All Controls']

    def get_all_control_families(self) -> Dict[str, CISControlFamily]:
        return self._control_families

    def get_all_control_domains_weight(self):
        control_domains = Counter([control.domain for control in self._cache['All Controls']])
        total = sum(control_domains.values())
        percentages = {key: round((value / total) * 100, 2) for key, value in control_domains.items()}
        return percentages

    def __repr__(self) -> str:
        return f'CISControlManager(workbook_path="{self._workbook_path}", config_path="{self._config_path}")'
