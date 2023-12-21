import json
import subprocess
from enum import Enum
import openpyxl
from DataModels import Recommendation, RecommendHeader, AuditCmd
from config_management.interfaces import IConfigLoader
from config_management.loaders import JSONConfigLoader
from workbook_management.workbook_manager import ExcelOpenWorkbook, ExcelValidator
from CISAuditManager import AuditCommandManager
from CISControlsManager import CISControlsProcessWorkbook
import re
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple, Set, List, Iterator, Generator
from utils.validation_utils import validate_and_return_file_path
from workbook_management.interfaces import IWorkbookLoader
from config_management.config_manager import BenchmarksConfigAttrs
from workbook_management.loaders import OpenPyXLWorkbookLoader


class CISBenchmarksConst(Enum):
    CIS_BENCHMARKS_CONFIG = 'CISBenchmarksConfig'
    RECOMMENDATION = 'recommendation'
    RECOMMEND_HEADER = 'recommend_header'
    ALLOWED_ITEMS = {RECOMMENDATION, RECOMMEND_HEADER}


class CISBenchmarksLoadConfig(BenchmarksConfigAttrs):
    def __init__(self, *, config_path: str, config_loader: IConfigLoader):
        self._config_path = validate_and_return_file_path(config_path, 'json')
        self._config_title = CISBenchmarksConst.CIS_BENCHMARKS_CONFIG.value
        super().__init__(config_loader)

    def _load_config(self) -> dict:
        config = self._config_loader.load(self._config_path).get(self._config_title)
        if not config:
            raise KeyError('The key does not exist within the configuration file.')
        return config

    @property
    def allowed_scope_levels(self) -> dict:
        scope_levels = {int(level): title for level, title in self._config.get('ALLOWED_SCOPE_LEVELS').items()}
        if not scope_levels:
            raise KeyError('The key does not exist within the configuration file.')
        return scope_levels

    @property
    def allowed_assessment_methods(self) -> list:
        allowed_assessment_methods = self._config.get('ALLOWED_ASSESSMENT_METHODS')
        if not allowed_assessment_methods:
            raise KeyError('The key does not exist within the configuration file.')
        return allowed_assessment_methods

    @property
    def benchmark_profiles_rex(self) -> str:
        benchmark_profiles_rex = self._config.get('BENCHMARK_PROFILES_REX')
        if not benchmark_profiles_rex:
            raise KeyError('The key does not exist within the configuration file.')
        return benchmark_profiles_rex

    @property
    def section(self) -> str:
        section = self._config.get('SECTION')
        if not section:
            raise KeyError('The key does not exist within the configuration file.')
        return section

    @property
    def recommendation(self) -> str:
        recommendation = self._config.get('RECOMMENDATION')
        if not recommendation:
            raise KeyError('The key does not exist within the configuration file.')
        return recommendation

    @property
    def title(self) -> str:
        title = self._config.get('TITLE')
        if not title:
            raise KeyError('The key does not exist within the configuration file.')
        return title

    @property
    def assessment_status(self) -> str:
        assessment_status = self._config.get('ASSESSMENT_STATUS')
        if not assessment_status:
            raise KeyError('The key does not exist within the configuration file.')
        return assessment_status

    @property
    def description(self) -> str:
        description = self._config.get('DESCRIPTION')
        if not description:
            raise KeyError('The key does not exist within the configuration file.')
        return description

    @property
    def rationale(self) -> str:
        rationale = self._config.get('RATIONALE')
        if not rationale:
            raise KeyError('The key does not exist within the configuration file.')
        return rationale

    @property
    def impact(self) -> str:
        impact = self._config.get('IMPACT')
        if not impact:
            raise KeyError('The key does not exist within the configuration file.')
        return impact

    @property
    def safeguard(self) -> str:
        safeguard = self._config.get('SAFEGUARD')
        if not safeguard:
            raise KeyError('The key does not exist within the configuration file.')
        return safeguard

    @property
    def overview_sheet(self) -> str:
        overview_sheet = self._config.get('OVERVIEW_SHEET')
        if not overview_sheet:
            raise KeyError('The key does not exist within the configuration file.')
        return overview_sheet

    @property
    def required_columns(self) -> set:
        required_column_titles = set(self._config.get('REQUIRED_COLUMN_TITLES'))
        if not required_column_titles:
            raise KeyError('The key does not exist within the configuration file.')
        return required_column_titles

    @property
    def workbooks_os_mapping(self) -> dict:
        workbooks_os_mapping = self._config.get('WORKBOOKS_OS_MAPPING')
        if not workbooks_os_mapping:
            raise KeyError('The key does not exist within the configuration file.')
        return workbooks_os_mapping

    @property
    def os_version_rex(self) -> str:
        os_version_rex = self._config.get('OS_VERSION_REX')
        if not os_version_rex:
            raise KeyError('The key does not exist within the configuration file.')
        return os_version_rex

    @property
    def os_versions_mapping(self) -> dict:
        os_versions_mapping = self._config.get('OS_VERSIONS_MAPPING')
        if not os_versions_mapping:
            raise KeyError('The key does not exist within the configuration file.')
        return os_versions_mapping

    def __repr__(self):
        return f'CISBenchmarksLoadConfig(config_path="{self._config_path}", config_loader="{self._config_loader}")'


class CISBenchmarksLoadWorkbook(ExcelOpenWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str):
        self._workbook_path = validate_and_return_file_path(workbook_path, 'xlsx')
        super().__init__(workbook_loader)

    def _load_workbook(self):
        return self._workbook_loader.load(self._workbook_path)


class CISBenchmarksWorkbookValidator(ExcelValidator):
    def __init__(self, workbook):
        super().__init__(workbook)

    @staticmethod
    def validate_column_titles(column_indices: dict, required_columns: set) -> bool:
        columns_to_check = column_indices.keys()
        if not required_columns.issubset(columns_to_check):
            missing_columns = required_columns.difference(columns_to_check)
            raise AttributeError(f"The following columns do not exist in the worksheet: '{', '.join(missing_columns)}'.")
        return True

    def validate_and_return_sheet_name(self, sheet_name: str) -> str:
        sheetnames_list = self._workbook.sheetnames
        if sheet_name not in sheetnames_list:
            raise ValueError(f'"{sheet_name}" is not in the sheet names. Possible sheet names: {sheetnames_list}.')
        return sheet_name

    @staticmethod
    def validate_and_return_scope_level(scope_level: int, allowed_scope_levels: set) -> int:
        if not isinstance(scope_level, int):
            raise TypeError(f'scope_level must be an integer, got {type(scope_level).__name__}')
        if scope_level not in allowed_scope_levels:
            raise ValueError(f'{scope_level} is not in the scope levels.')
        return scope_level

    def validate_and_return_benchmark_scope_profile(self, scope_level: int, scope_levels_os_mapping: dict, allowed_scope_levels: set) -> str:
        scope_level = self.validate_and_return_scope_level(scope_level, allowed_scope_levels)
        scope_level_os = scope_levels_os_mapping.get(scope_level)
        if scope_level_os is None:
            raise ValueError(f'Benchmark profile for level {scope_level} does not exist.')
        return scope_level_os

    @staticmethod
    def validate_and_return_item_id(item_id: str) -> str:
        if not isinstance(item_id, str):
            raise TypeError(f'item_id must be a string, got {type(item_id).__name__}')
        return item_id

    @staticmethod
    def validate_assessment_method_type(assessment_method: str, allowed_assessment_methods: list) -> str:
        if assessment_method.casefold() not in allowed_assessment_methods:
            raise ValueError(f"{assessment_method} is not in allowed assessment methods. The allowed assessment methods are: '{allowed_assessment_methods}'.")
        return assessment_method


class CISBenchmarksProcessWorkbook(CISBenchmarksLoadWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str = None, benchmarks_config: CISBenchmarksLoadConfig):
        self._config = benchmarks_config
        if workbook_path is None:
            workbook_path = self._get_os_version_workbook_path()
        super().__init__(workbook_loader=workbook_loader, workbook_path=workbook_path)
        self._validator = CISBenchmarksWorkbookValidator(self._workbook)
        self._scope_levels_os_mapping = self._get_scope_levels_os_mapping()
        self._allowed_scope_levels = set(map(int, self._config.allowed_scope_levels.keys()))
        self._recommendations_cache = {}
        self._headers_cache = {}
        self._populate_benchmark_cache_and_headers()

    def _get_current_os_version(self) -> str:
        os_version_rex = self._config.os_version_rex
        os_versions_mapping = self._config.os_versions_mapping
        allowed_os_versions = set(os_versions_mapping.values())

        try:
            os_cmd = subprocess.run('sw_vers', stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
            stdout = os_cmd.stdout.decode('UTF-8').split('\n')
            stderr = os_cmd.stderr.decode('UTF-8').split('\n')
            return_code = os_cmd.returncode

            if return_code != 0:
                return stderr[0]

            match = re.findall(os_version_rex, stdout[1])
            if not match:
                raise ValueError(f"OS version regex match failed. Regex pattern: '{os_version_rex}'")

            rex_os = match[0]
            os_version = os_versions_mapping.get(rex_os)

            if not os_version:
                raise ValueError(f'"{os_version}" does not exist.')
            if os_version not in allowed_os_versions:
                raise KeyError(f"'{os_version}' is not in the allowed OS versions. Allowed OS versions: '{', '.join(allowed_os_versions)}'")

            return os_version

        except (RuntimeError, ValueError, IndexError, KeyError) as error:
            print(f"Error occurred: '{error}'.")

    def _get_os_version_workbook_path(self):
        os_version = self._get_current_os_version()
        workbooks_os_mapping = self._config.workbooks_os_mapping
        os_version_workbook_path = workbooks_os_mapping.get(os_version)
        if not os_version_workbook_path:
            raise ValueError(f'OS version path for {os_version_workbook_path} does not exist.')
        return os_version_workbook_path

    def _get_overview_worksheet(self) -> Worksheet:
        sheet_name = self._validator.validate_and_return_sheet_name(self._config.overview_sheet)
        overview_worksheet = self._workbook[sheet_name]
        if not overview_worksheet:
            raise KeyError(f'"{sheet_name}" sheet name cannot be found.')
        return overview_worksheet

    def _get_scope_levels_os_mapping(self) -> Dict:
        overview_worksheet = self._get_overview_worksheet()
        regex_pattern = self._config.benchmark_profiles_rex
        scope_levels_os_mapping = {}
        for row in overview_worksheet.iter_rows(values_only=True):
            for cell in row:
                match = re.search(regex_pattern, str(cell))
                if match:
                    try:
                        os_system, level = match.groups()
                        scope_levels_os_mapping[int(level)] = os_system
                    except (IndexError, ValueError) as e:
                        raise ValueError(f'Invalid data format in cell: {cell}. Error: {e}')
        return scope_levels_os_mapping

    def _get_worksheet_scope_headers(self, scope_level: int) -> Tuple[Worksheet, Dict[str, int]]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        curr_sheet_level = self._config.allowed_scope_levels[scope_level]
        sheet_name = self._validator.validate_and_return_sheet_name(curr_sheet_level)
        worksheet = self._workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {title: index for index, title in enumerate(header_row)}
        return worksheet, column_indices

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[Tuple[str, str, str, bool]]:
        if self._validator.validate_column_titles(column_indices, self._config.required_columns):
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                recommend_id = row[column_indices[self._config.recommendation]]
                title = row[column_indices[self._config.title]]
                description = row[column_indices[self._config.description]]
                rationale = row[column_indices[self._config.rationale]]
                impact = row[column_indices[self._config.impact]]
                safeguard_id = row[column_indices[self._config.safeguard]]
                assessment_method = row[column_indices[self._config.assessment_status]]
                is_header = False

                if not assessment_method:
                    is_header = True
                    recommend_id = row[column_indices[self._config.section]]

                yield recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header

    def _get_worksheet_all_scopes_row_attributes(self) -> Generator:
        all_scopes_mapping = self._scope_levels_os_mapping.items()
        for level, benchmark_profile in all_scopes_mapping:
            worksheet, column_indices = self._get_worksheet_scope_headers(level)
            worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
            yield level, benchmark_profile, worksheet_row_attrs

    def _initialize_cache_and_headers_keys(self) -> Tuple[Dict[str, List], Dict[str, List]]:
        cache_mapping, headers_mapping = {}, {}
        for _, profile in self._scope_levels_os_mapping.items():
            cache_mapping[profile], headers_mapping[profile] = [], []
        return cache_mapping, headers_mapping

    def _populate_benchmark_cache_and_headers(self):
        self._recommendations_cache, self._headers_cache = self._initialize_cache_and_headers_keys()
        all_scopes_attributes = self._get_worksheet_all_scopes_row_attributes()
        for level, profile, worksheet_row_attrs in all_scopes_attributes:
            for recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header in worksheet_row_attrs:
                if is_header:
                    header = RecommendHeader(recommend_id=recommend_id, level=level, title=title, description=description)
                    self._headers_cache[profile].append(header)
                else:
                    recommendation = Recommendation(recommend_id=recommend_id, level=level, title=title,
                                                    rationale=rationale,
                                                    impact=impact, safeguard_id=safeguard_id,
                                                    assessment_method=assessment_method)
                    self._recommendations_cache[profile].append(recommendation)

    def _get_item_by_id(self, item_id: str, cache: dict, scope_profile: str):
        item_id = self._validator.validate_and_return_item_id(item_id)
        scope_items = cache.get(scope_profile)

        if scope_items is None:
            raise KeyError(f'Scope items for level "{scope_profile}" cannot be found.')

        for item in scope_items:
            if item_id == item.recommend_id:
                return item
        raise KeyError(f'Item with ID "{item_id}" is not in level "{scope_profile}".')

    def get_recommendation_by_id(self, *, scope_level: int = 1, recommendation_id: str) -> Recommendation:
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level, self._scope_levels_os_mapping, self._allowed_scope_levels)
        return self._get_item_by_id(recommendation_id, self._recommendations_cache, scope_profile)

    def get_recommendation_header_by_id(self, *, scope_level: int = 1, header_id: str) -> Recommendation:
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level, self._scope_levels_os_mapping, self._allowed_scope_levels)
        return self._get_item_by_id(header_id, self._headers_cache, scope_profile)

    def get_all_levels_recommendations(self) -> Dict[str, List[Dict[str, Recommendation]]]:
        if not self._recommendations_cache:
            raise KeyError('Cache is empty.')
        return self._recommendations_cache

    def get_all_levels_recommendation_headers(self) -> Dict[str, List[RecommendHeader]]:
        if not self._headers_cache:
            raise KeyError('Headers cache is empty.')
        return self._headers_cache

    def get_recommendations_by_level(self, *, scope_level: int = 1) -> List[Recommendation]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level, self._scope_levels_os_mapping, self._allowed_scope_levels)
        if scope_profile not in self._recommendations_cache:
            raise KeyError(f'"{scope_profile}" scope profile is not in the cache.')
        return self._recommendations_cache.get(scope_profile)

    def get_recommendation_headers_by_level(self, *, scope_level: int = 1) -> List[Recommendation]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level, self._scope_levels_os_mapping, self._allowed_scope_levels)
        if scope_profile not in self._headers_cache:
            raise KeyError(f'"{scope_profile}" scope profile is not in the cache.')
        return self._headers_cache.get(scope_profile)

    def get_recommendations_by_assessment_method(self, *, scope_level: int = 1, assessment_method: str = None) -> Generator:
        assessment_method = self._validator.validate_assessment_method_type(assessment_method, self._config.allowed_assessment_methods)
        recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
        for recommendation in recommendations_scope:
            if assessment_method == recommendation.assessment_method.casefold():
                yield recommendation


json_config_loader = JSONConfigLoader()
openpyxl_workbook_loader = OpenPyXLWorkbookLoader()

cis_benchmarks_config = CISBenchmarksLoadConfig(config_loader=json_config_loader,
                                                config_path='config/cis_workbooks_config.json')
workbook_processor = CISBenchmarksProcessWorkbook(workbook_loader=openpyxl_workbook_loader,
                                                  workbook_path='cis_benchmarks/CIS_Apple_macOS_14.0_Sonoma_Benchmark_v1.0.0.xlsx',
                                                  benchmarks_config=cis_benchmarks_config)

print(workbook_processor.get_all_levels_recommendations())
print(workbook_processor.get_all_levels_recommendation_headers())
print(workbook_processor.get_recommendation_headers_by_level(scope_level=2))
print(workbook_processor.get_recommendation_header_by_id(scope_level=1, header_id='1'))
print(workbook_processor.get_recommendations_by_level(scope_level=2))
print(workbook_processor.get_recommendations_by_level())
print(list(workbook_processor.get_recommendations_by_assessment_method(assessment_method='manual')))
print(workbook_processor.get_recommendation_by_id(scope_level=2, recommendation_id='2.1.1.1'))


#
# print(cis_benchmarks_config.required_columns)
# print(cis_benchmarks_workbook._workbook.sheetnames)


# class CISBenchmarkManager(ExcelOpenWorkbook):
#     def __init__(self, *, workbook_path: str, audit_manager: AuditCommandManager, cis_control_manager: CISControlsProcessWorkbook):
#         super().__init__(workbook_path)
#         self._workbook = self.load_workbook()
#         self._config = self.load_config()
#         self._audit_manager = audit_manager
#         self._cis_control_manager = cis_control_manager
#         self._benchmark_profiles = self._get_benchmark_profiles()
#         self._scope_levels_os_mapping = self._populate_scope_levels_os_mapping()
#         self._headers = None
#         self._populate_benchmark_cache_and_headers()
#         self._map_recommendations_and_cis_controls()
#         self._map_recommendations_and_audit_commands()
#
#     def load_workbook(self):
#         return openpyxl.load_workbook(self.workbook_path)
#
#     def load_config(self):
#         try:
#             with open(self.config_path, 'r') as config_file:
#                 return json.load(config_file)[__class__.__name__]
#         except json.JSONDecodeError as e:
#             raise ValueError(f'Error parsing JSON file at {self.config_path}: {e}')
#
#     @property
#     def benchmark_profiles(self) -> List[Tuple]:
#         return self._benchmark_profiles
#
#     @property
#     def scope_levels_os_mapping(self) -> Dict:
#         return self._scope_levels_os_mapping
#
#     @property
#     def scope_levels(self) -> Dict:
#         return {int(level): title for level, title in self._config['SCOPE_LEVELS'].items()}
#
#     @property
#     def allowed_assessment_methods(self) -> Set:
#         return self._config['ALLOWED_ASSESSMENT_METHODS']
#
#     @property
#     def benchmark_profiles_rex(self) -> str:
#         return self._config['BENCHMARK_PROFILES_REX']
#
#     @property
#     def recommendation(self) -> str:
#         return self._config['RECOMMENDATION']
#
#     @property
#     def rationale(self) -> str:
#         return self._config['RATIONALE']
#
#     @property
#     def impact(self) -> str:
#         return self._config['IMPACT']
#
#     @property
#     def assess_status(self) -> str:
#         return self._config['ASSESS_STATUS']
#
#     @property
#     def section(self) -> str:
#         return self._config['SECTION']
#
#     @property
#     def overview_sheet(self) -> str:
#         return self._config['OVERVIEW_SHEET']
#
#     def _get_benchmark_profiles(self) -> list:
#         regex_pattern = self.benchmark_profiles_rex
#         sheet_name = self._validate_and_return_sheet_name(self.overview_sheet)
#         overview_worksheet = self._workbook[sheet_name]
#         overview_paragraphs = str([paragraph for paragraph in overview_worksheet.iter_rows(values_only=True)])
#         return re.findall(regex_pattern, overview_paragraphs)
#
#     def _populate_scope_levels_os_mapping(self) -> Dict:
#         benchmark_profiles = self._benchmark_profiles
#         scope_levels_os_mapping = defaultdict(list)
#         for profile, profile_level in benchmark_profiles:
#             scope_levels_os_mapping[int(profile_level)].append(profile)
#         return scope_levels_os_mapping
#
#     def _validate_and_return_benchmark_scope_profile(self, scope_level: int) -> str:
#         scope_level = self._validate_and_return_scope_level(scope_level)
#         scope_level_os = self._scope_levels_os_mapping[scope_level]
#         if not scope_level_os:
#             raise ValueError(f'Benchmark profile for level {scope_level} does not exist.')
#         return next(iter(scope_level_os))
#
#     def _validate_and_return_scope_level(self, scope_level: int) -> int:
#         if not isinstance(scope_level, int):
#             raise TypeError(f'scope_level must be an integer, got {type(scope_level).__name__}')
#         if scope_level not in self.scope_levels:
#             raise ValueError(f'{scope_level} is not in the scope levels.')
#         return scope_level
#
#     @staticmethod
#     def _validate_and_return_item_id(item_id: str) -> str:
#         if not isinstance(item_id, str):
#             raise TypeError(f'item_id must be a string, got {type(item_id).__name__}')
#         return item_id
#
#     def _validate_and_get_items_by_type(self, scope_level: int, item_type: str) -> List[Recommendation] | List[RecommendHeader]:
#         if item_type.casefold() == 'recommendation':
#             scope_items = self.get_recommendations_by_level(scope_level=scope_level)
#         elif item_type.casefold() == 'recommend_header':
#             scope_items = self.get_recommendations_scope_headers(scope_level=scope_level)
#         else:
#             raise KeyError(
#                 f'Invalid item type "{item_type}" provided. Item types can be either "recommendation" or "recommend_header".')
#         return scope_items
#
#     def _validate_assessment_method_type(self, assessment_method: str) -> str:
#         if assessment_method is None:
#             raise ValueError("Assessment method cannot be 'None'.")
#         if assessment_method.casefold() not in self.allowed_assessment_methods:
#             raise ValueError(
#                 f"{assessment_method} is not in allowed assessment methods. The allowed assessment methods are: '{self.allowed_assessment_methods}'.")
#         return assessment_method
#
#     def _get_worksheet_scope_headers(self, scope_level: int) -> Tuple[Worksheet, Dict[str, int]]:
#         scope_level = self._validate_and_return_scope_level(scope_level)
#         curr_sheet_level = self.scope_levels[scope_level]
#         sheet_name = self._validate_and_return_sheet_name(curr_sheet_level)
#         worksheet = self._workbook[sheet_name]
#         header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
#         column_indices = {title: index for index, title in enumerate(header_row)}
#
#         return worksheet, column_indices
#
#     def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[Tuple[str, str, str, bool]]:
#         if self._validate_column_titles(column_indices, self.required_column_titles):
#             for row in worksheet.iter_rows(min_row=2, values_only=True):
#                 recommend_id = row[column_indices[self.recommendation]]
#                 title = row[column_indices[self.title]]
#                 description = row[column_indices[self.description]]
#                 rationale = row[column_indices[self.rationale]]
#                 impact = row[column_indices[self.impact]]
#                 safeguard_id = row[column_indices[self.safeguard]]
#                 assessment_method = row[column_indices[self.assess_status]]
#                 is_header = False
#
#                 if not assessment_method:
#                     is_header = True
#                     recommend_id = row[column_indices[self.section]]
#
#                 yield recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header
#
#     def _get_worksheet_all_scopes_row_attributes(self) -> Generator:
#         all_scopes_mapping = self._scope_levels_os_mapping.items()
#         for level, benchmark_profiles in all_scopes_mapping:
#             worksheet, column_indices = self._get_worksheet_scope_headers(level)
#             worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
#             for profile in benchmark_profiles:
#                 yield level, profile, worksheet_row_attrs
#
#     def _initialize_cache_and_headers_keys(self) -> Tuple[Dict[str, List], Dict[str, List]]:
#         cache_mapping = {}
#         headers_mapping = {}
#         for _, benchmark_profiles in self._scope_levels_os_mapping.items():
#             for profile in benchmark_profiles:
#                 cache_mapping[profile] = []
#                 headers_mapping[profile] = []
#         return cache_mapping, headers_mapping
#
#     def _populate_benchmark_cache_and_headers(self):
#         self._cache, self._headers = self._initialize_cache_and_headers_keys()
#         all_scopes_attributes = self._get_worksheet_all_scopes_row_attributes()
#         for level, profile, worksheet_row_attrs in all_scopes_attributes:
#             for recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header in worksheet_row_attrs:
#                 if is_header:
#                     header = RecommendHeader(recommend_id=recommend_id, level=level, title=title,
#                                              description=description)
#                     self._headers[profile].append(header)
#                 else:
#                     recommendation = Recommendation(recommend_id=recommend_id, level=level, title=title,
#                                                     rationale=rationale,
#                                                     impact=impact, safeguard_id=safeguard_id,
#                                                     assessment_method=assessment_method)
#                     self._cache[profile].append(recommendation)
#
#     def get_item_by_id(self, *, scope_level: int = 1, item_id: str, item_type: str) -> Recommendation | RecommendHeader:
#         scope_level = self._validate_and_return_scope_level(scope_level)
#         item_id = self._validate_and_return_item_id(item_id)
#         scope_items = self._validate_and_get_items_by_type(scope_level, item_type)
#
#         for item in scope_items:
#             if item_id == item.recommend_id:
#                 return item
#
#         raise KeyError(f'{item_type.capitalize()} with ID {item_id} is not in level {scope_level} of {item_type}s.')
#
#     def get_recommendations_by_level(self, *, scope_level: int = 1) -> List[Recommendation]:
#         scope_level = self._validate_and_return_scope_level(scope_level)
#         scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
#         return self._cache[scope_profile]
#
#     def get_all_levels_recommendations(self) -> Dict[str, List[Dict[str, Recommendation]]]:
#         return self._cache
#
#     def get_recommendations_scope_headers(self, *, scope_level: int = 1) -> List[RecommendHeader]:
#         scope_level = self._validate_and_return_scope_level(scope_level)
#         scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
#         return self._headers[scope_profile]
#
#     def get_all_scopes_recommendation_headers(self) -> Dict[str, List[RecommendHeader]]:
#         return self._headers
#
#     def get_recommendations_by_assessment_method(self, *, scope_level: int = 1, assessment_method: str = None) -> Generator:
#         assessment_method = self._validate_assessment_method_type(assessment_method)
#         recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
#         for recommendation in recommendations_scope:
#             if assessment_method == recommendation.assessment_method.casefold():
#                 yield recommendation
#
#     def _map_recommendations_and_cis_controls(self):
#         all_cis_controls = {control.safeguard_id: control for control in self._cis_control_manager.get_all_controls()}
#         for scope_level in self.scope_levels:
#             recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
#             for recommendation in recommendations_scope:
#                 control = all_cis_controls.get(recommendation.safeguard_id)
#                 if control:
#                     recommendation.cis_control = control
#
#     def _map_recommendations_and_audit_commands(self):
#         audit_commands = self._audit_manager.audit_commands
#         commands_map = {cmd['recommend_id']: cmd for cmd in audit_commands}
#
#         for scope_level in self.scope_levels:
#             recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
#             for recommendation in recommendations_scope:
#                 if recommendation.recommend_id in commands_map:
#                     cmd = commands_map[recommendation.recommend_id]
#                     command, expected_output = self._audit_manager.get_command_attrs(cmd)
#                     recommendation.audit_cmd = AuditCmd(command=command, expected_output=expected_output)
#
#     def evaluate_recommendations_compliance(self, *, scope_level: int = 1) -> List[Recommendation]:
#         recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
#         for recommendation in recommendations_scope:
#             audit_cmd = recommendation.audit_cmd
#             if audit_cmd:
#                 command = audit_cmd.command
#                 expected_output = audit_cmd.expected_output
#                 audit_result = self._audit_manager.run_command(command, expected_output)
#                 recommendation.compliant = audit_result
#                 yield recommendation
#
#     def __repr__(self) -> str:
#         return f'CISBenchmarkManager(workbook_path="{self.workbook_path}", config_path="{self.config_path}")'
