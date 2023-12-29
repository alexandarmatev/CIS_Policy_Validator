import subprocess
from collections import namedtuple
from enum import Enum
from cis_audit_manager import CISAuditLoadCommands
from data_models.data_models import Recommendation, RecommendHeader
from config_management.interfaces import IConfigLoader
from workbook_management.workbook_manager import ExcelOpenWorkbook, ExcelValidator
import re
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple, Set, List, Iterator, Generator
from utils.validation_utils import validate_and_return_file_path
from workbook_management.interfaces import IWorkbookLoader
from config_management.config_manager import BenchmarksConfigAttrs


class CISBenchmarksConst(Enum):
    CIS_BENCHMARKS_CONFIG = 'CISBenchmarksConfig'
    RECOMMENDATION = 'recommendation'
    RECOMMEND_HEADER = 'recommend_header'
    ALLOWED_ITEMS = {RECOMMENDATION, RECOMMEND_HEADER}


class CISBenchmarksLoadConfig(BenchmarksConfigAttrs):
    def __init__(self, *, config_path: str, config_loader: IConfigLoader):
        self._config_path = validate_and_return_file_path(config_path, 'json')
        self._config_title = CISBenchmarksConst.CIS_BENCHMARKS_CONFIG.value
        super().__init__(config_loader)

    def _load_config(self) -> Dict:
        config = self._config_loader.load(self._config_path).get(self._config_title)
        if not config:
            raise KeyError('The key does not exist within the configuration file.')
        return config

    @property
    def allowed_scope_levels(self) -> Dict:
        scope_levels = {int(level): title for level, title in self._config.get('ALLOWED_SCOPE_LEVELS').items()}
        if not scope_levels:
            raise KeyError('The key does not exist within the configuration file.')
        return scope_levels

    @property
    def allowed_assessment_methods(self) -> List:
        allowed_assessment_methods = self._config.get('ALLOWED_ASSESSMENT_METHODS')
        if not allowed_assessment_methods:
            raise KeyError('The key does not exist within the configuration file.')
        return allowed_assessment_methods

    @property
    def benchmark_profiles_rex(self) -> str:
        benchmark_profiles_rex = self._config.get('BENCHMARK_PROFILES_REX')
        if not benchmark_profiles_rex:
            raise KeyError('The key does not exist within the configuration file.')
        return benchmark_profiles_rex

    @property
    def section(self) -> str:
        section = self._config.get('SECTION')
        if not section:
            raise KeyError('The key does not exist within the configuration file.')
        return section

    @property
    def recommendation(self) -> str:
        recommendation = self._config.get('RECOMMENDATION')
        if not recommendation:
            raise KeyError('The key does not exist within the configuration file.')
        return recommendation

    @property
    def title(self) -> str:
        title = self._config.get('TITLE')
        if not title:
            raise KeyError('The key does not exist within the configuration file.')
        return title

    @property
    def assessment_status(self) -> str:
        assessment_status = self._config.get('ASSESSMENT_STATUS')
        if not assessment_status:
            raise KeyError('The key does not exist within the configuration file.')
        return assessment_status

    @property
    def description(self) -> str:
        description = self._config.get('DESCRIPTION')
        if not description:
            raise KeyError('The key does not exist within the configuration file.')
        return description

    @property
    def rationale(self) -> str:
        rationale = self._config.get('RATIONALE')
        if not rationale:
            raise KeyError('The key does not exist within the configuration file.')
        return rationale

    @property
    def impact(self) -> str:
        impact = self._config.get('IMPACT')
        if not impact:
            raise KeyError('The key does not exist within the configuration file.')
        return impact

    @property
    def safeguard(self) -> str:
        safeguard = self._config.get('SAFEGUARD')
        if not safeguard:
            raise KeyError('The key does not exist within the configuration file.')
        return safeguard

    @property
    def overview_sheet(self) -> str:
        overview_sheet = self._config.get('OVERVIEW_SHEET')
        if not overview_sheet:
            raise KeyError('The key does not exist within the configuration file.')
        return overview_sheet

    @property
    def required_columns(self) -> Set:
        required_column_titles = set(self._config.get('REQUIRED_COLUMN_TITLES'))
        if not required_column_titles:
            raise KeyError('The key does not exist within the configuration file.')
        return required_column_titles

    @property
    def workbooks_os_mapping(self) -> Dict:
        workbooks_os_mapping = self._config.get('WORKBOOKS_OS_MAPPING')
        if not workbooks_os_mapping:
            raise KeyError('The key does not exist within the configuration file.')
        return workbooks_os_mapping

    @property
    def os_version_rex(self) -> str:
        os_version_rex = self._config.get('OS_VERSION_REX')
        if not os_version_rex:
            raise KeyError('The key does not exist within the configuration file.')
        return os_version_rex

    @property
    def custom_os_version_rex(self) -> str:
        custom_os_version_rex = self._config.get('CUSTOM_OS_VERSION_REX')
        if not custom_os_version_rex:
            raise KeyError('The key does not exist within the configuration file.')
        return custom_os_version_rex

    @property
    def os_versions_mapping(self) -> Dict:
        os_versions_mapping = self._config.get('OS_VERSIONS_MAPPING')
        if not os_versions_mapping:
            raise KeyError('The key does not exist within the configuration file.')
        return os_versions_mapping

    def __repr__(self):
        return f'CISBenchmarksLoadConfig(config_path="{self._config_path}", config_loader="{self._config_loader}")'


class CISBenchmarksLoadWorkbook(ExcelOpenWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str):
        self._workbook_path = validate_and_return_file_path(workbook_path, 'xlsx')
        super().__init__(workbook_loader)

    def _load_workbook(self):
        return self._workbook_loader.load(self._workbook_path)


class CISBenchmarksWorkbookValidator(ExcelValidator):
    def __init__(self, workbook):
        super().__init__(workbook)

    @staticmethod
    def validate_column_titles(column_indices: Dict, required_columns: Set) -> bool:
        columns_to_check = column_indices.keys()
        if not required_columns.issubset(columns_to_check):
            missing_columns = required_columns.difference(columns_to_check)
            raise AttributeError(
                f"The following columns do not exist in the worksheet: '{', '.join(missing_columns)}'.")
        return True

    def validate_and_return_sheet_name(self, sheet_name: str) -> str:
        sheetnames_list = self._workbook.sheetnames
        if sheet_name not in sheetnames_list:
            raise ValueError(f'"{sheet_name}" is not in the sheet names. Possible sheet names: {sheetnames_list}.')
        return sheet_name

    @staticmethod
    def validate_and_return_scope_level(scope_level: int, allowed_scope_levels: Set) -> int:
        if not isinstance(scope_level, int):
            raise TypeError(f'scope_level must be an integer, got {type(scope_level).__name__}')
        if scope_level not in allowed_scope_levels:
            raise ValueError(f'{scope_level} is not in the scope levels.')
        return scope_level

    def validate_and_return_benchmark_scope_profile(self, scope_level: int, scope_levels_os_mapping: Dict,
                                                    allowed_scope_levels: Set) -> str:
        scope_level = self.validate_and_return_scope_level(scope_level, allowed_scope_levels)
        scope_level_os = scope_levels_os_mapping.get(scope_level)
        if scope_level_os is None:
            raise ValueError(f'Benchmark profile for level {scope_level} does not exist.')
        return scope_level_os

    @staticmethod
    def validate_and_return_item_id(item_id: str) -> str:
        if not isinstance(item_id, str):
            raise TypeError(f'item_id must be a string, got {type(item_id).__name__}')
        return item_id

    @staticmethod
    def validate_assessment_method_type(assessment_method: str, allowed_assessment_methods: List) -> str:
        if assessment_method.casefold() not in allowed_assessment_methods:
            raise ValueError(
                f"{assessment_method} is not in allowed assessment methods. The allowed assessment methods are: '{allowed_assessment_methods}'.")
        return assessment_method


class CISBenchmarksProcessWorkbook(CISBenchmarksLoadWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str = None,
                 benchmarks_config: CISBenchmarksLoadConfig, cis_controls: List, commands_loader: CISAuditLoadCommands):
        self._config = benchmarks_config
        if workbook_path is None:
            workbook_path = self._get_os_version_workbook_path()
            self._audit_commands = commands_loader.get_os_specific_commands(self._get_current_os_version())
        else:
            self._audit_commands = commands_loader.get_os_specific_commands(self._get_custom_os_version(workbook_path))
        super().__init__(workbook_loader=workbook_loader, workbook_path=workbook_path)
        self._validator = CISBenchmarksWorkbookValidator(self._workbook)
        self._cis_controls = cis_controls
        self._scope_levels_os_mapping = self._get_scope_levels_os_mapping()
        self._allowed_scope_levels = set(map(int, self._config.allowed_scope_levels.keys()))
        self._recommendations_cache = {}
        self._headers_cache = {}
        self._populate_benchmark_cache_and_headers()
        self._map_recommendations_and_audit_commands()
        self._map_recommendations_and_cis_controls()

    def _get_current_os_version(self) -> str:
        os_version_rex = self._config.os_version_rex
        os_versions_mapping = self._config.os_versions_mapping
        allowed_os_versions = set(os_versions_mapping.values())

        try:
            os_cmd = subprocess.run('sw_vers', stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
            stdout = os_cmd.stdout.decode('UTF-8').split('\n')
            stderr = os_cmd.stderr.decode('UTF-8').split('\n')
            return_code = os_cmd.returncode

            if return_code != 0:
                return stderr[0]

            match = re.findall(os_version_rex, stdout[1])
            if not match:
                raise ValueError(f"OS version regex match failed. Regex pattern: '{os_version_rex}'")

            rex_os = match[0]
            os_version = os_versions_mapping.get(rex_os)

            if not os_version:
                raise ValueError(f'"{os_version}" does not exist.')
            if os_version not in allowed_os_versions:
                raise KeyError(
                    f"'{os_version}' is not in the allowed OS versions. Allowed OS versions: '{', '.join(allowed_os_versions)}'")

            return os_version

        except (RuntimeError, ValueError, IndexError, KeyError) as error:
            print(f"Error occurred: '{error}'.")

    def _get_os_version_workbook_path(self):
        os_version = self._get_current_os_version()
        workbooks_os_mapping = self._config.workbooks_os_mapping
        os_version_workbook_path = workbooks_os_mapping.get(os_version)
        if not os_version_workbook_path:
            raise ValueError(f'OS version path for {os_version_workbook_path} does not exist.')
        return os_version_workbook_path

    def _get_custom_os_version(self, workbook_path: str):
        custom_os_version_rex = self._config.custom_os_version_rex
        regex_result = re.search(custom_os_version_rex, workbook_path).group(1)
        custom_os_version = self._config.os_versions_mapping.get(regex_result)
        if not custom_os_version:
            raise ValueError('OS version cannot be found.')
        return custom_os_version

    def _get_overview_worksheet(self) -> Worksheet:
        sheet_name = self._validator.validate_and_return_sheet_name(self._config.overview_sheet)
        overview_worksheet = self._workbook[sheet_name]
        if not overview_worksheet:
            raise KeyError(f'"{sheet_name}" sheet name cannot be found.')
        return overview_worksheet

    def _get_scope_levels_os_mapping(self) -> Dict:
        overview_worksheet = self._get_overview_worksheet()
        regex_pattern = self._config.benchmark_profiles_rex
        scope_levels_os_mapping = {}
        for row in overview_worksheet.iter_rows(values_only=True):
            for cell in row:
                match = re.search(regex_pattern, str(cell))
                if match:
                    try:
                        os_system, level = match.groups()
                        scope_levels_os_mapping[int(level)] = os_system
                    except (IndexError, ValueError) as e:
                        raise ValueError(f'Invalid data format in cell: {cell}. Error: {e}')
        return scope_levels_os_mapping

    def _get_worksheet_scope_headers(self, scope_level: int) -> Tuple[Worksheet, Dict[str, int]]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        curr_sheet_level = self._config.allowed_scope_levels[scope_level]
        sheet_name = self._validator.validate_and_return_sheet_name(curr_sheet_level)
        worksheet = self._workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {title: index for index, title in enumerate(header_row)}
        return worksheet, column_indices

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[
        Tuple[str, str, str, bool]]:
        if self._validator.validate_column_titles(column_indices, self._config.required_columns):
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                recommend_id = row[column_indices[self._config.recommendation]]
                title = row[column_indices[self._config.title]]
                description = row[column_indices[self._config.description]]
                rationale = row[column_indices[self._config.rationale]]
                impact = row[column_indices[self._config.impact]]
                safeguard_id = row[column_indices[self._config.safeguard]]
                assessment_method = row[column_indices[self._config.assessment_status]]
                is_header = False

                if not assessment_method:
                    is_header = True
                    recommend_id = row[column_indices[self._config.section]]

                yield recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header

    def _get_worksheet_all_scopes_row_attributes(self) -> Generator:
        all_scopes_mapping = self._scope_levels_os_mapping.items()
        for level, benchmark_profile in all_scopes_mapping:
            worksheet, column_indices = self._get_worksheet_scope_headers(level)
            worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
            yield level, benchmark_profile, worksheet_row_attrs

    def _initialize_cache_and_headers_keys(self) -> Tuple[Dict[str, List], Dict[str, List]]:
        cache_mapping, headers_mapping = {}, {}
        for _, profile in self._scope_levels_os_mapping.items():
            cache_mapping[profile], headers_mapping[profile] = [], []
        return cache_mapping, headers_mapping

    def _populate_benchmark_cache_and_headers(self):
        self._recommendations_cache, self._headers_cache = self._initialize_cache_and_headers_keys()
        all_scopes_attributes = self._get_worksheet_all_scopes_row_attributes()
        for level, profile, worksheet_row_attrs in all_scopes_attributes:
            for recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header in worksheet_row_attrs:
                if is_header:
                    header = RecommendHeader(recommend_id=recommend_id, level=level, title=title,
                                             description=description)
                    self._headers_cache[profile].append(header)
                else:
                    recommendation = Recommendation(recommend_id=recommend_id, level=level, title=title,
                                                    rationale=rationale,
                                                    impact=impact, safeguard_id=safeguard_id,
                                                    assessment_method=assessment_method)
                    self._recommendations_cache[profile].append(recommendation)

    def _get_item_by_id(self, item_id: str, cache: Dict, scope_profile: str):
        item_id = self._validator.validate_and_return_item_id(item_id)
        scope_items = cache.get(scope_profile)

        if scope_items is None:
            raise KeyError(f'Scope items for level "{scope_profile}" cannot be found.')

        for item in scope_items:
            if item_id == item.recommend_id:
                return item
        raise KeyError(f'Item with ID "{item_id}" is not in level "{scope_profile}".')

    def _map_recommendations_and_audit_commands(self):
        AuditCmd = namedtuple('AuditCmd', ['recommend_id', 'title', 'command', 'expected_output'])
        audit_commands = self._audit_commands
        commands_map = {cmd['recommend_id']: AuditCmd(**cmd) for cmd in audit_commands}
        for level in self._allowed_scope_levels:
            recommendations = self.get_recommendations_by_level(scope_level=level)
            for recommendation in recommendations:
                if recommendation.recommend_id in commands_map:
                    recommendation.audit_cmd = commands_map[recommendation.recommend_id]

    def _map_recommendations_and_cis_controls(self):
        all_cis_controls = {control.safeguard_id: control for control in self._cis_controls}
        for level in self._allowed_scope_levels:
            recommendations = self.get_recommendations_by_level(scope_level=level)
            for recommendation in recommendations:
                control = all_cis_controls.get(recommendation.safeguard_id)
                if control:
                    recommendation.cis_control = control

    def get_recommendation_by_id(self, *, scope_level: int = 1, recommendation_id: str) -> Recommendation:
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level,
                                                                                    self._scope_levels_os_mapping,
                                                                                    self._allowed_scope_levels)
        return self._get_item_by_id(recommendation_id, self._recommendations_cache, scope_profile)

    def get_recommendation_header_by_id(self, *, scope_level: int = 1, header_id: str) -> Recommendation:
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level,
                                                                                    self._scope_levels_os_mapping,
                                                                                    self._allowed_scope_levels)
        return self._get_item_by_id(header_id, self._headers_cache, scope_profile)

    def get_all_levels_recommendations(self) -> List[Recommendation]:
        all_levels_recommendations = []
        for level in self._allowed_scope_levels:
            all_levels_recommendations.extend(self.get_recommendations_by_level(scope_level=level))
        if not all_levels_recommendations:
            raise KeyError('No recommendations have been found.')
        return all_levels_recommendations

    def get_all_levels_recommendation_headers(self) -> List[RecommendHeader]:
        all_levels_headers = []
        for level in self._allowed_scope_levels:
            all_levels_headers.extend(self.get_recommendation_headers_by_level(scope_level=level))
        if not all_levels_headers:
            raise KeyError('No recommendation headers have been found.')
        return all_levels_headers

    def get_recommendations_by_level(self, *, scope_level: int = 1) -> List[Recommendation]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level,
                                                                                    self._scope_levels_os_mapping,
                                                                                    self._allowed_scope_levels)
        if scope_profile not in self._recommendations_cache:
            raise KeyError(f'"{scope_profile}" scope profile is not in the cache.')
        return self._recommendations_cache.get(scope_profile)

    def get_recommendation_headers_by_level(self, *, scope_level: int = 1) -> List[RecommendHeader]:
        scope_level = self._validator.validate_and_return_scope_level(scope_level, self._allowed_scope_levels)
        scope_profile = self._validator.validate_and_return_benchmark_scope_profile(scope_level,
                                                                                    self._scope_levels_os_mapping,
                                                                                    self._allowed_scope_levels)
        if scope_profile not in self._headers_cache:
            raise KeyError(f'"{scope_profile}" scope profile is not in the cache.')
        return self._headers_cache.get(scope_profile)

    def get_recommendations_by_assessment_method(self, *, scope_level: int = 1,
                                                 assessment_method: str = None) -> Generator:
        assessment_method = self._validator.validate_assessment_method_type(assessment_method,
                                                                            self._config.allowed_assessment_methods)
        recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
        for recommendation in recommendations_scope:
            if assessment_method == recommendation.assessment_method.casefold():
                yield recommendation
