from abc import ABC, abstractmethod
from typing import Set, Dict, Iterator, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from workbook_management.interfaces import IWorkbookLoader, IConfigLoader


class ExcelOpenWorkbook(ABC):
    def __init__(self, workbook_loader: IWorkbookLoader, config_loader: IConfigLoader):
        self._workbook_loader = workbook_loader
        self._config_loader = config_loader
        self._workbook = self.load_workbook()
        self._config = self.load_config()

    @abstractmethod
    def load_workbook(self):
        pass

    @abstractmethod
    def load_config(self):
        pass

    @property
    def safeguard(self) -> str:
        return self._config['SAFEGUARD']

    @property
    def title(self) -> str:
        return self._config['TITLE']

    @property
    def description(self) -> str:
        return self._config['DESCRIPTION']

    @property
    def required_column_titles(self) -> Set:
        return set(self._config['REQUIRED_COLUMN_TITLES'])

    @staticmethod
    def _validate_column_titles(column_indices: dict, required_columns: Set[str]) -> bool:
        columns_to_check = column_indices.keys()
        if not required_columns.issubset(columns_to_check):
            missing_columns = required_columns.difference(columns_to_check)
            raise AttributeError(
                f"The following columns do not exist in the worksheet: '{', '.join(missing_columns)}'.")
        return True

    def _validate_and_return_sheet_name(self, sheet_name: str) -> str:
        sheetnames_list = self._workbook.sheetnames
        if sheet_name not in sheetnames_list:
            raise ValueError(f'"{sheet_name}" is not in the sheet names. Possible sheet names: {sheetnames_list}.')
        return sheet_name
