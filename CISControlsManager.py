from typing import Tuple, Dict, NamedTuple, List
from openpyxl.worksheet.worksheet import Worksheet
from workbook_management.workbook_manager import ExcelOpenWorkbook
from config_management.config_manager import ControlsConfigAttrs
from DataModels import CISControl, CISControlFamily
from collections import namedtuple
from collections import Counter
from workbook_management.interfaces import IWorkbookLoader
from workbook_management.workbook_manager import ExcelValidator
from config_management.interfaces import IConfigLoader
from utils.validation_utils import validate_and_return_file_path
from enum import Enum


class CISControlsConst(Enum):
    CIS_CONTROLS_CONFIG = 'CISControlsConfig'


class CISControlsLoadConfig(ControlsConfigAttrs):
    def __init__(self, *, config_path: str, config_loader: IConfigLoader):
        self._config_path = validate_and_return_file_path(config_path, 'json')
        self._config_title = CISControlsConst.CIS_CONTROLS_CONFIG.value
        super().__init__(config_loader)

    def _load_config(self) -> dict:
        config = self._config_loader.load(self._config_path).get(self._config_title)
        if not config:
            raise KeyError('This configuration does not exist within the configuration file.')
        return config

    @property
    def worksheet_name(self) -> str:
        ws_name = self._config.get('WORKSHEET_NAME')
        if ws_name:
            return ws_name
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def cis_safeguard(self) -> str:
        cis_safeguard = self._config.get('SAFEGUARD')
        if cis_safeguard:
            return cis_safeguard
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def control_family_id(self) -> str:
        control_family_id = self._config.get('CONTROL_FAMILY_ID')
        if control_family_id:
            return control_family_id
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def asset_type(self) -> str:
        asset_type = self._config.get('ASSET_TYPE')
        if asset_type:
            return asset_type
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def domain(self) -> str:
        domain = self._config.get('DOMAIN')
        if domain:
            return domain
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def title(self) -> str:
        title = self._config.get('TITLE')
        if title:
            return title
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def description(self) -> str:
        description = self._config.get('DESCRIPTION')
        if description:
            return description
        raise KeyError('The key does not exist within the configuration file.')

    @property
    def required_columns(self) -> set:
        required_columns = set(self._config.get('REQUIRED_COLUMN_TITLES'))
        if required_columns:
            return required_columns
        raise KeyError('The key does not exist within the configuration file.')

    def __repr__(self):
        return f'CISControlsLoadConfig(config_path="{self._config_path}", config_loader="{self._config_loader}")'


class CISControlsLoadWorkbook(ExcelOpenWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str):
        self._workbook_path = validate_and_return_file_path(workbook_path, 'xlsx')
        super().__init__(workbook_loader)

    def _load_workbook(self):
        return self._workbook_loader.load(self._workbook_path)


class CISControlsWorkbookValidator(ExcelValidator):
    def __init__(self, workbook):
        super().__init__(workbook)

    @staticmethod
    def validate_column_titles(column_indices: dict, required_columns: set) -> bool:
        columns_to_check = column_indices.keys()
        if not required_columns.issubset(columns_to_check):
            missing_columns = required_columns.difference(columns_to_check)
            raise AttributeError(
                f"The following columns do not exist in the worksheet: '{', '.join(missing_columns)}'.")
        return True

    def validate_and_return_sheet_name(self, sheet_name: str) -> str:
        sheetnames_list = self._workbook.sheetnames
        if sheet_name not in sheetnames_list:
            raise ValueError(f'"{sheet_name}" is not in the sheet names. Possible sheet names: {sheetnames_list}.')
        return sheet_name


class CISControlsProcessWorkbook(CISControlsLoadWorkbook):
    def __init__(self, *, workbook_loader: IWorkbookLoader, workbook_path: str, controls_config: CISControlsLoadConfig):
        super().__init__(workbook_loader=workbook_loader, workbook_path=workbook_path)
        self._config = controls_config
        self._excel_validator = CISControlsWorkbookValidator(self._workbook)
        self._cache = {'All Controls': []}
        self._control_families = {}
        self._populate_controls_cache()

    def _get_worksheet_scope_headers(self) -> Tuple[Worksheet, Dict[str, int]]:
        worksheet_name = self._excel_validator.validate_and_return_sheet_name(self._config.worksheet_name)
        worksheet = self._workbook[worksheet_name]
        if worksheet:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            column_indices = {title: index for index, title in enumerate(header_row)}
            return worksheet, column_indices
        raise KeyError(f'"{worksheet}" worksheet cannot be found.')

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> NamedTuple:
        RowData = namedtuple('RowData', ['safeguard_id', 'asset_type', 'domain', 'title', 'description', 'control_family_id', 'is_family'])
        required_columns = self._config.required_columns
        if self._excel_validator.validate_column_titles(column_indices, required_columns):
            safeguard_ids = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                safeguard_id = str(row[column_indices[self._config.cis_safeguard]])
                if safeguard_id in safeguard_ids:
                    safeguard_id += '0'
                safeguard_ids.add(safeguard_id)
                asset_type = row[column_indices[self._config.asset_type]]
                domain = row[column_indices[self._config.domain]]
                title = row[column_indices[self._config.title]]
                description = row[column_indices[self._config.description]]
                control_family_id = None
                is_family = False

                if not asset_type:
                    is_family = True
                    control_family_id = str(row[column_indices[self._config.control_family_id]])

                yield RowData(safeguard_id, asset_type, domain, title, description, control_family_id, is_family)

    def _populate_controls_cache(self):
        worksheet, column_indices = self._get_worksheet_scope_headers()
        worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
        for row_data in worksheet_row_attrs:
            if row_data.is_family:
                control_family_id = row_data.control_family_id.strip()
                self._control_families[control_family_id] = CISControlFamily(title=row_data.title,
                                                                             description=row_data.description)
            else:
                self._cache['All Controls'].append(
                    CISControl(safeguard_id=row_data.safeguard_id, asset_type=row_data.asset_type,
                               domain=row_data.domain, title=row_data.title, description=row_data.description))

    def get_all_controls(self) -> List[CISControl]:
        return self._cache['All Controls']

    def get_all_control_families(self) -> Dict[str, CISControlFamily]:
        return self._control_families

    def get_all_control_domains_weight(self):
        control_domains = Counter([control.domain for control in self._cache['All Controls']])
        total = sum(control_domains.values())
        percentages = {key: round((value / total) * 100, 2) for key, value in control_domains.items()}
        return percentages



