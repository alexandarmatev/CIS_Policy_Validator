from DataModels import Recommendation, RecommendHeader, AuditCmd
from ExcelWorkbookBase import ExcelWorkbookBase
from AuditCommandManager import AuditCommandManager
from CISControlManager import CISControlManager
import re
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple, Set, List, Iterator, Generator


class CISBenchmarkManager(ExcelWorkbookBase):
    """
    Manages and processes CIS (Center for Internet Security) benchmarks within an Excel workbook.
    Inherits from ExcelWorkbookBase and utilizes an AuditCommandManager instance to determine
    the correct workbook path based on the OS system version. Assumes a constant configuration path.

    Attributes inherited from ExcelWorkbookBase:
        _workbook_path (str): Path to the Excel workbook file.
        _workbook (openpyxl.Workbook): Instance of the openpyxl Workbook.
        _config_path (str): Path to the JSON configuration file.
        _config (dict): Configuration data specific to the class instance.
        _cache (dict): Cache for storing processed data.
    """
    def __init__(self, *, workbook_path: str, config_path: str, audit_manager: AuditCommandManager, cis_control_manager: CISControlManager):
        """
        Initializes the CISBenchmarkManager with an audit manager which determines the appropriate
        workbook path based on the current OS system version. Utilizes a constant configuration path.

        Parameters:
            workbook_path: The path to the Excel workbook.
            config_path: The path to the JSON configuration file.
            audit_manager: An instance of AuditCommandManager responsible for managing audit commands and determining the correct workbook path.
            cis_control_manager: An instance of CISControlManager responsible for managing and accessing CIS control data.
        """
        super().__init__(workbook_path, config_path)
        self._audit_manager = audit_manager
        self._cis_control_manager = cis_control_manager
        self._benchmark_profiles = self._get_benchmark_profiles()
        self._scope_levels_os_mapping = self._populate_scope_levels_os_mapping()
        self._headers = None
        self._populate_benchmark_cache_and_headers()
        self._map_recommendations_and_cis_controls()
        self._map_recommendations_and_audit_commands()

    @property
    def benchmark_profiles(self) -> List[Tuple]:
        """
        Gets the extracted benchmark profiles.

        Returns:
            A list of tuples representing the benchmark profiles.
        """
        return self._benchmark_profiles

    @property
    def scope_levels_os_mapping(self) -> Dict:
        """
        Gets the mapping of scope levels to operating systems.

        Returns:
            A dictionary mapping scope levels to lists of operating systems.
        """
        return self._scope_levels_os_mapping

    @property
    def scope_levels(self) -> Dict:
        """
        Gets the scope levels from the configuration.

        Returns:
            A dictionary mapping scope level integers to their titles.
        """
        return {int(level): title for level, title in self.config['SCOPE_LEVELS'].items()}

    @property
    def allowed_assessment_methods(self) -> Set:
        """
        Gets the allowed assessment methods from the configuration.

        Returns:
            A set of allowed assessment methods.
        """
        return self.config['ALLOWED_ASSESSMENT_METHODS']

    @property
    def benchmark_profiles_rex(self) -> str:
        """
        Gets the regular expression used for extracting benchmark profiles.

        Returns:
            A string representing the regular expression for benchmark profiles.
        """
        return self.config['BENCHMARK_PROFILES_REX']

    @property
    def recommendation(self) -> str:
        """
        Gets the recommendation key from the configuration.

        Returns:
            Recommendation key.
        """
        return self._config['RECOMMENDATION']

    @property
    def rationale(self) -> str:
        """
        Gets the rationale key from the configuration.

        Returns:
            Rationale key.
        """
        return self.config['RATIONALE']

    @property
    def impact(self) -> str:
        """
        Gets the impact key from the configuration.

        Returns:
            Impact key.
        """
        return self.config['IMPACT']

    @property
    def assess_status(self) -> str:
        """
        Gets the assessment status key from the configuration.

        Returns:
            Assessment status key.
        """
        return self.config['ASSESS_STATUS']

    @property
    def section(self) -> str:
        """
        Gets the section key from the configuration.

        Returns:
            Section key.
        """
        return self.config['SECTION']

    @property
    def overview_sheet(self) -> str:
        """
        Gets the name of the overview sheet from the configuration.

        Returns:
            Name of the overview sheet.
        """
        return self.config['OVERVIEW_SHEET']

    def _get_benchmark_profiles(self) -> list:
        """
        Extracts benchmark profiles using a regular expression from the overview sheet of the workbook.

        Returns:
            A list of tuples containing benchmark profiles and their respective levels.
        """
        regex_pattern = self.benchmark_profiles_rex
        sheet_name = self._validate_and_return_sheet_name(self.overview_sheet)
        overview_worksheet = self._workbook[sheet_name]
        overview_paragraphs = str([paragraph for paragraph in overview_worksheet.iter_rows(values_only=True)])
        return re.findall(regex_pattern, overview_paragraphs)

    def _populate_scope_levels_os_mapping(self) -> Dict:
        """
        Creates a mapping of scope levels to operating systems using benchmark profiles.

        Returns:
            A dictionary where keys are scope levels and values are lists of operating systems.
        """
        benchmark_profiles = self._benchmark_profiles
        scope_levels_os_mapping = defaultdict(list)
        for profile, profile_level in benchmark_profiles:
            scope_levels_os_mapping[int(profile_level)].append(profile)
        return scope_levels_os_mapping

    def _validate_and_return_benchmark_scope_profile(self, scope_level: int) -> str:
        """
        Validates the given scope level and returns the first operating system associated with it.

        Parameters:
            scope_level: An integer representing the scope level.

        Returns:
            str: The name of the first operating system associated with the given scope level.

        Raises:
            ValueError: If there is no benchmark profile for the given scope level.
        """
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_level_os = self._scope_levels_os_mapping[scope_level]
        if not scope_level_os:
            raise ValueError(f'Benchmark profile for level {scope_level} does not exist.')
        return next(iter(scope_level_os))

    def _validate_and_return_scope_level(self, scope_level: int) -> int:
        """
        Validates that the given scope level is an integer and exists within the defined scope levels.

        Parameters:
            scope_level: An integer representing the scope level.

        Returns:
            The validated scope level.

        Raises:
            TypeError: If the scope level is not an integer.
            ValueError: If the scope level is not in the defined scope levels.
        """
        if not isinstance(scope_level, int):
            raise TypeError(f'scope_level must be an integer, got {type(scope_level).__name__}')
        if scope_level not in self.scope_levels:
            raise ValueError(f'{scope_level} is not in the scope levels.')
        return scope_level

    @staticmethod
    def _validate_and_return_item_id(item_id: str) -> str:
        """
        Validates that the given item ID is a string.

        Parameters:
            item_id: A string representing the item ID.

        Returns:
            The validated item ID.

        Raises:
            TypeError: If the item ID is not a string.
        """
        if not isinstance(item_id, str):
            raise TypeError(f'item_id must be a string, got {type(item_id).__name__}')
        return item_id

    def _validate_and_get_items_by_type(self, scope_level: int, item_type: str) -> List[Recommendation] | List[RecommendHeader]:
        """
        Validates the item type and returns a list of items (either Recommendations or RecommendHeaders) for the given scope level.

        Parameters:
            scope_level: An integer representing the scope level.
            item_type: A string representing the type of items to retrieve ('recommendation' or 'recommend_header').

        Returns:
            A list of either Recommendation or RecommendHeader objects for the given scope level.

        Raises:
            KeyError: If an invalid item type is provided.
        """
        if item_type.casefold() == 'recommendation':
            scope_items = self.get_recommendations_by_level(scope_level=scope_level)
        elif item_type.casefold() == 'recommend_header':
            scope_items = self.get_recommendations_scope_headers(scope_level=scope_level)
        else:
            raise KeyError(
                f'Invalid item type "{item_type}" provided. Item types can be either "recommendation" or "recommend_header".')
        return scope_items

    def _validate_assessment_method_type(self, assessment_method: str) -> str:
        """
        Validates the assessment method against the allowed assessment methods.

        Parameters:
            assessment_method: A string representing the assessment method to validate.

        Returns:
            The validated assessment method.

        Raises:
            ValueError: If the assessment method is None or not in the allowed assessment methods.
        """
        if assessment_method is None:
            raise ValueError("Assessment method cannot be 'None'.")
        if assessment_method.casefold() not in self.allowed_assessment_methods:
            raise ValueError(
                f"{assessment_method} is not in allowed assessment methods. The allowed assessment methods are: '{self.allowed_assessment_methods}'.")
        return assessment_method

    def _get_worksheet_scope_headers(self, scope_level: int) -> Tuple[Worksheet, Dict[str, int]]:
        """
        Retrieves the worksheet and header column indices for the given scope level.

        Parameters:
            scope_level: An integer representing the scope level.

        Returns:
            A tuple containing the worksheet and a dictionary of column headers with their respective indices.
        """
        scope_level = self._validate_and_return_scope_level(scope_level)
        curr_sheet_level = self.scope_levels[scope_level]
        sheet_name = self._validate_and_return_sheet_name(curr_sheet_level)
        worksheet = self._workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {title: index for index, title in enumerate(header_row)}

        return worksheet, column_indices

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[Tuple[str, str, str, bool]]:
        """
        Iterates through worksheet rows, extracting attributes based on the provided column indices.

        Parameters:
            worksheet: The worksheet to iterate through.
            column_indices: A dictionary mapping column titles to their respective indices.

        Returns:
            An iterator that yields tuples containing extracted row attributes.
        """
        if self._validate_column_titles(column_indices, self.required_column_titles):
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                recommend_id = row[column_indices[self.recommendation]]
                title = row[column_indices[self.title]]
                description = row[column_indices[self.description]]
                rationale = row[column_indices[self.rationale]]
                impact = row[column_indices[self.impact]]
                safeguard_id = row[column_indices[self.safeguard]]
                assessment_method = row[column_indices[self.assess_status]]
                is_header = False

                if not assessment_method:
                    is_header = True
                    recommend_id = row[column_indices[self.section]]

                yield recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header

    def _get_worksheet_all_scopes_row_attributes(self) -> Generator:
        """
        Generates worksheet row attributes for all scope levels and their corresponding benchmark profiles.

        Returns:
            A generator that yields tuples containing scope level, profile, and worksheet row attributes.
        """
        all_scopes_mapping = self._scope_levels_os_mapping.items()
        for level, benchmark_profiles in all_scopes_mapping:
            worksheet, column_indices = self._get_worksheet_scope_headers(level)
            worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
            for profile in benchmark_profiles:
                yield level, profile, worksheet_row_attrs

    def _initialize_cache_and_headers_keys(self) -> Tuple[Dict[str, List], Dict[str, List]]:
        """
        Initializes the cache and headers with keys based on benchmark profiles.

        Returns:
            Two dictionaries for cache and headers, respectively, with benchmark profiles as keys.
        """
        cache_mapping = {}
        headers_mapping = {}
        for _, benchmark_profiles in self._scope_levels_os_mapping.items():
            for profile in benchmark_profiles:
                cache_mapping[profile] = []
                headers_mapping[profile] = []
        return cache_mapping, headers_mapping

    def _populate_benchmark_cache_and_headers(self):
        """
        Populates the benchmark cache and headers with recommendations and headers extracted from the worksheets.
        """
        self._cache, self._headers = self._initialize_cache_and_headers_keys()
        all_scopes_attributes = self._get_worksheet_all_scopes_row_attributes()
        for level, profile, worksheet_row_attrs in all_scopes_attributes:
            for recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header in worksheet_row_attrs:
                if is_header:
                    header = RecommendHeader(recommend_id=recommend_id, level=level, title=title,
                                             description=description)
                    self._headers[profile].append(header)
                else:
                    recommendation = Recommendation(recommend_id=recommend_id, level=level, title=title,
                                                    rationale=rationale,
                                                    impact=impact, safeguard_id=safeguard_id,
                                                    assessment_method=assessment_method)
                    self._cache[profile].append(recommendation)

    def get_item_by_id(self, *, scope_level: int = 1, item_id: str, item_type: str) -> Recommendation | RecommendHeader:
        """
        Retrieves an item (either Recommendation or RecommendHeader) by its ID for a given scope level and item type.

        Parameters:
            scope_level: An optional integer representing the scope level (default is 1).
            item_id: A string representing the item ID.
            item_type: A string representing the type of item ('recommendation' or 'recommend_header').

        Returns:
            The requested item.

        Raises:
            KeyError: If the item with the given ID is not found in the specified level and type.
        """
        scope_level = self._validate_and_return_scope_level(scope_level)
        item_id = self._validate_and_return_item_id(item_id)
        scope_items = self._validate_and_get_items_by_type(scope_level, item_type)

        for item in scope_items:
            if item_id == item.recommend_id:
                return item

        raise KeyError(f'{item_type.capitalize()} with ID {item_id} is not in level {scope_level} of {item_type}s.')

    def get_recommendations_by_level(self, *, scope_level: int = 1) -> List[Recommendation]:
        """
        Retrieves a list of recommendations for a specified scope level.

        Parameters:
            scope_level: An optional integer representing the scope level (default is 1).

        Returns:
            A list of Recommendation objects for the specified scope level.
        """
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
        return self._cache[scope_profile]

    def get_all_levels_recommendations(self) -> Dict[str, List[Dict[str, Recommendation]]]:
        """
        Retrieves recommendations for all scope levels.

        Returns:
            A dictionary where keys are scope levels and values are lists of Recommendation objects.
        """
        return self._cache

    def get_recommendations_scope_headers(self, *, scope_level: int = 1) -> List[RecommendHeader]:
        """
        Retrieves a list of recommendation headers for a specified scope level.

        Parameters:
            scope_level: An optional integer representing the scope level (default is 1).

        Returns:
            A list of RecommendHeader objects for the specified scope level.
        """
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
        return self._headers[scope_profile]

    def get_all_scopes_recommendation_headers(self) -> Dict[str, List[RecommendHeader]]:
        """
        Retrieves recommendation headers for all scope levels.

        Returns:
            A dictionary where keys are scope levels and values are lists of RecommendHeader objects.
        """
        return self._headers

    def get_recommendations_by_assessment_method(self, *, scope_level: int = 1, assessment_method: str = None) -> Generator:
        """
        Generates recommendations for a specified scope level and assessment method.

        Parameters:
            scope_level: An optional integer representing the scope level (default is 1).
            assessment_method: An optional string representing the assessment method.

        Returns:
            Recommendation: Yields Recommendation objects that match the specified assessment method.
        """
        assessment_method = self._validate_assessment_method_type(assessment_method)
        recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
        for recommendation in recommendations_scope:
            if assessment_method == recommendation.assessment_method.casefold():
                yield recommendation

    def _map_recommendations_and_cis_controls(self):
        """
        Maps CIS controls to recommendations within the specified scope.
        """
        all_cis_controls = {control.safeguard_id: control for control in self._cis_control_manager.get_all_controls()}
        for scope_level in self.scope_levels:
            recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
            for recommendation in recommendations_scope:
                control = all_cis_controls.get(recommendation.safeguard_id)
                if control:
                    recommendation.cis_control = control

    def _map_recommendations_and_audit_commands(self):
        """
        Maps audit commands to recommendations for a specified scope level and operating system version.
        """
        audit_commands = self._audit_manager.audit_commands
        commands_map = {cmd['recommend_id']: cmd for cmd in audit_commands}

        for scope_level in self.scope_levels:
            recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
            for recommendation in recommendations_scope:
                if recommendation.recommend_id in commands_map:
                    cmd = commands_map[recommendation.recommend_id]
                    command, expected_output = self._audit_manager.get_command_attrs(cmd)
                    recommendation.audit_cmd = AuditCmd(command=command, expected_output=expected_output)

    def evaluate_recommendations_compliance(self, *, scope_level: int = 1) -> List[Recommendation]:
        """
        Evaluates the compliance of recommendations based on audit commands and CIS controls for a specified scope level and operating system version.

        Parameters:
            scope_level: An optional integer representing the scope level (default is 1).

        Returns:
            Yields Recommendation objects after evaluating their compliance.
        """
        recommendations_scope = self.get_recommendations_by_level(scope_level=scope_level)
        for recommendation in recommendations_scope:
            audit_cmd = recommendation.audit_cmd
            if audit_cmd:
                command = audit_cmd.command
                expected_output = audit_cmd.expected_output
                audit_result = self._audit_manager.run_command(command, expected_output)
                recommendation.compliant = audit_result
                yield recommendation

    def __repr__(self) -> str:
        """
        Represents the CISBenchmarkManager instance as a string.

        Returns:
            A string representation of the CISBenchmarkManager instance, including paths to the workbook and configuration file.
        """
        return f'CISBenchmarkManager(workbook_path="{self.workbook_path}", config_path="{self.config_path}")'



