from DataModels import Recommendation, RecommendHeader
from ExcelWorkbookBase import ExcelWorkbookBase
from utils.file_utils import validate_and_return_file_path
import openpyxl
import re
from collections import defaultdict
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Tuple, Set, List, Iterator, Generator


class WorkbookManager(ExcelWorkbookBase):
    _SCOPE_LEVELS = {1: 'Level 1', 2: 'Level 2'}
    _ALLOWED_ASSESSMENT_METHODS = {'manual', 'automated'}
    _BENCHMARK_PROFILES_REX = r'(Level\s(\d)\s-\s\w+ \d+\.\d{1,2}(?:\s\w+)*)'
    _SECTION, _RECOMMENDATION, _TITLE, _ASSESS_STATS, _DESCR, _RATIONALE, _IMPACT, _SAFEGUARD = 'Section #', 'Recommendation #', 'Title', 'Assessment Status', \
                                                                                                'Description', 'Rationale Statement', 'Impact Statement', 'CIS Safeguards 1 (v8)'
    _REQUIRED_COLUMN_TITLES = {_SECTION, _RECOMMENDATION, _TITLE, _ASSESS_STATS, _RATIONALE, _IMPACT, _SAFEGUARD}

    def __init__(self, workbook_path: str):
        super().__init__(workbook_path)
        self._scope_levels = self.get_scope_levels()
        self._benchmark_profiles_rex = self.get_benchmark_profiles_rex()
        self._benchmark_profiles = self._get_benchmark_profiles()
        self._scope_levels_os_mapping = self._populate_scope_levels_os_mapping()
        self._cache = None
        self._headers = None
        self._populate_cache_and_headers()

    @property
    def path(self) -> str:
        return self._workbook_path

    @path.setter
    def path(self, new_path: str):
        self._workbook_path = validate_and_return_file_path(new_path)
        self._workbook = openpyxl.load_workbook(self._workbook_path)
        self._benchmark_profiles = self._get_benchmark_profiles()
        self._scope_levels_os_mapping = self._populate_scope_levels_os_mapping()
        self._cache = None
        self._headers = None
        self._populate_cache_and_headers()

    @property
    def benchmark_profiles(self) -> List[Tuple]:
        return self._get_benchmark_profiles()

    @property
    def scope_levels_os_mapping(self) -> Dict:
        return self._scope_levels_os_mapping

    @classmethod
    def get_scope_levels(cls) -> Dict:
        return cls._SCOPE_LEVELS

    @classmethod
    def get_allowed_assessment_methods(cls) -> Set:
        return cls._ALLOWED_ASSESSMENT_METHODS

    @classmethod
    def get_benchmark_profiles_rex(cls) -> str:
        return cls._BENCHMARK_PROFILES_REX

    @classmethod
    def get_required_column_titles(cls) -> Set:
        return cls._REQUIRED_COLUMN_TITLES

    def _get_benchmark_profiles(self) -> list:
        regex_pattern = self._benchmark_profiles_rex
        sheet_name = self._validate_and_return_sheet_name('Overview - Glossary')
        overview_worksheet = self._workbook[sheet_name]
        overview_paragraphs = str([paragraph for paragraph in overview_worksheet.iter_rows(values_only=True)])
        return re.findall(regex_pattern, overview_paragraphs)

    def _populate_scope_levels_os_mapping(self) -> Dict:
        benchmark_profiles = self._benchmark_profiles
        scope_levels_os_mapping = defaultdict(list)
        for profile, profile_level in benchmark_profiles:
            scope_levels_os_mapping[int(profile_level)].append(profile)
        return scope_levels_os_mapping

    def _validate_and_return_benchmark_scope_profile(self, scope_level: int) -> str:
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_level_os = self._scope_levels_os_mapping[scope_level]
        if not scope_level_os:
            raise ValueError(f'Benchmark profile for level {scope_level} does not exist.')
        return next(iter(scope_level_os))

    def _validate_and_return_scope_level(self, scope_level: int) -> int:
        if not isinstance(scope_level, int):
            raise TypeError(f'scope_level must be an integer, got {type(scope_level).__name__}')
        if scope_level not in self._scope_levels:
            raise ValueError(f'{scope_level} is not in the scope levels.')
        return scope_level

    @staticmethod
    def _validate_and_return_item_id(item_id: str) -> str:
        if not isinstance(item_id, str):
            raise TypeError(f'item_id must be a string, got {type(item_id).__name__}')
        return item_id

    def _validate_and_get_items_by_type(self, scope_level: int, item_type: str) -> List[Dict[str, Recommendation]] | List[Dict[str, RecommendHeader]]:
        if item_type.casefold() == 'recommendation':
            scope_items = self.get_recommendations_scope(scope_level=scope_level)
        elif item_type.casefold() == 'recommend_header':
            scope_items = self.get_recommendations_scope_headers(scope_level=scope_level)
        else:
            raise KeyError(f'Invalid item type "{item_type}" provided. Item types can be either "recommendation" or "recommend_header".')
        return scope_items

    def _validate_assessment_method_type(self, assessment_method: str) -> str:
        allowed_assessment_methods = self.get_allowed_assessment_methods()
        if assessment_method is None:
            raise ValueError("Assessment method cannot be 'None'.")
        if assessment_method.casefold() not in allowed_assessment_methods:
            raise ValueError(f"{assessment_method} is not in allowed assessment methods. The allowed assessment methods are: '{allowed_assessment_methods}'.")
        return assessment_method

    def _get_worksheet_scope_headers(self, scope_level: int) -> Tuple[Worksheet, Dict[str, int]]:
        scope_level = self._validate_and_return_scope_level(scope_level)
        curr_sheet_level = self._scope_levels[scope_level]
        sheet_name = self._validate_and_return_sheet_name(curr_sheet_level)
        worksheet = self._workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_indices = {title: index for index, title in enumerate(header_row)}

        return worksheet, column_indices

    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[Tuple[str, str, str, bool]]:
        required_columns = self.get_required_column_titles()
        if self._validate_column_titles(column_indices, required_columns):
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                recommend_id = row[column_indices[WorkbookManager._RECOMMENDATION]]
                title = row[column_indices[WorkbookManager._TITLE]]
                description = row[column_indices[WorkbookManager._DESCR]]
                rationale = row[column_indices[WorkbookManager._RATIONALE]]
                impact = row[column_indices[WorkbookManager._IMPACT]]
                safeguard_id = row[column_indices[WorkbookManager._SAFEGUARD]]
                assessment_method = row[column_indices[WorkbookManager._ASSESS_STATS]]
                is_header = False

                if not assessment_method:
                    is_header = True
                    recommend_id = row[column_indices[WorkbookManager._SECTION]]

                yield recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header

    def _get_worksheet_all_scopes_row_attributes(self) -> Generator:
        all_scopes_mapping = self._scope_levels_os_mapping.items()
        for level, benchmark_profiles in all_scopes_mapping:
            worksheet, column_indices = self._get_worksheet_scope_headers(level)
            worksheet_row_attrs = self._get_worksheet_row_attributes(worksheet, column_indices)
            for profile in benchmark_profiles:
                yield level, profile, worksheet_row_attrs

    def _initialize_cache_and_headers_keys(self) -> Tuple[Dict[str, List], Dict[str, List]]:
        cache_mapping = {}
        headers_mapping = {}
        for _, benchmark_profiles in self._scope_levels_os_mapping.items():
            for profile in benchmark_profiles:
                cache_mapping[profile] = []
                headers_mapping[profile] = []
        return cache_mapping, headers_mapping

    def _populate_cache_and_headers(self):
        self._cache, self._headers = self._initialize_cache_and_headers_keys()
        all_scopes_attributes = self._get_worksheet_all_scopes_row_attributes()
        for level, profile, worksheet_row_attrs in all_scopes_attributes:
            for recommend_id, title, description, rationale, impact, safeguard_id, assessment_method, is_header in worksheet_row_attrs:
                if is_header:
                    header = RecommendHeader(header_id=recommend_id, level=level, title=title, description=description)
                    self._headers[profile].append({recommend_id: header})
                else:
                    recommendation = Recommendation(recommend_id=recommend_id, level=level, title=title, rationale=rationale,
                                                    impact=impact, safeguard_id=safeguard_id, assessment_method=assessment_method)
                    self._cache[profile].append({recommend_id: recommendation})

    def get_item_by_id(self, *, scope_level: int = 1, item_id: str, item_type: str) -> Recommendation | RecommendHeader:
        scope_level = self._validate_and_return_scope_level(scope_level)
        item_id = self._validate_and_return_item_id(item_id)
        scope_items = self._validate_and_get_items_by_type(scope_level, item_type)

        for item_dict in scope_items:
            if item_id in item_dict:
                return item_dict[item_id]

        raise KeyError(f'{item_type.capitalize()} with ID {item_id} is not in level {scope_level} of {item_type}s.')

    def get_recommendations_scope(self, *, scope_level: int = 1) -> List[Dict[str, Recommendation]]:
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
        return self._cache[scope_profile]

    def get_all_scopes_recommendations(self) -> Dict[str, List[Dict[str, Recommendation]]]:
        return self._cache

    def get_recommendations_scope_headers(self, *, scope_level: int = 1) -> List[Dict[str, RecommendHeader]]:
        scope_level = self._validate_and_return_scope_level(scope_level)
        scope_profile = self._validate_and_return_benchmark_scope_profile(scope_level)
        return self._headers[scope_profile]

    def get_all_scopes_recommendation_headers(self) -> Dict[str, List[Dict[str, RecommendHeader]]]:
        return self._headers

    def get_recommendations_by_assessment_method(self, *, scope_level: int = 1, assessment_method: str = None) -> Generator:
        assessment_method = self._validate_assessment_method_type(assessment_method)
        recommendations_scope = self.get_recommendations_scope(scope_level=scope_level)
        for dict_ in recommendations_scope:
            for id_, recommendation in dict_.items():
                if assessment_method == recommendation.assessment_method.casefold():
                    yield recommendation

    def __repr__(self):
        return f"WorkbookManager(workbook_path='{self._workbook_path}', workbook='{self._workbook}', " \
               f"scope_levels='{self._scope_levels}', benchmark_profiles_rex='{self._benchmark_profiles_rex}', " \
               f"scope_levels_os_mapping='{self._scope_levels_os_mapping}', benchmark_profiles='{self._benchmark_profiles}')"
