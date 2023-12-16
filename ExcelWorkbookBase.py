from typing import Set, Dict, Iterator, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from utils.validation_utils import validate_and_return_file_path
from abc import ABC, abstractmethod


class ExcelOpenWorkbook(ABC):
    def __init__(self, workbook_path: str):
        self._workbook_path = validate_and_return_file_path(workbook_path, 'xlsx')

    @abstractmethod
    def load_workbook(self):
        pass

    @abstractmethod
    def load_config(self):
        pass

    @property
    def config_path(self) -> str:
        return self._config_path

    @property
    def workbook_path(self) -> str:
        return self._workbook_path

    @property
    def safeguard(self) -> str:
        """
        Gets the safeguard value from the configuration.

        Returns:
            Safeguard value.
        """
        return self._config['SAFEGUARD']

    @property
    def title(self) -> str:
        """
        Gets the title value from the configuration.

        Returns:
            Title value.
        """
        return self._config['TITLE']

    @property
    def description(self) -> str:
        """
        Gets the description value from the configuration.

        Returns:
            Description value.
        """
        return self._config['DESCRIPTION']

    @property
    def required_column_titles(self) -> Set:
        """
        Gets the set of required column titles from the configuration.

        Returns:
            A set of required column titles.
        """
        return set(self._config['REQUIRED_COLUMN_TITLES'])

    @staticmethod
    def _validate_column_titles(column_indices: dict, required_columns: Set[str]) -> bool:
        """
        Validates if the required columns exist in the provided column indices.

        Parameters:
            column_indices: A dictionary mapping column titles to their respective indices.
            required_columns: A set of required column titles.

        Returns:
            True if validation passes, otherwise raises an AttributeError.

        Raises:
            AttributeError: If any required columns are missing in the column indices.
        """
        columns_to_check = column_indices.keys()
        if not required_columns.issubset(columns_to_check):
            missing_columns = required_columns.difference(columns_to_check)
            raise AttributeError(
                f"The following columns do not exist in the worksheet: '{', '.join(missing_columns)}'.")
        return True

    @abstractmethod
    def _get_worksheet_row_attributes(self, worksheet: Worksheet, column_indices: Dict[str, int]) -> Iterator[
        Tuple[str, str, str, bool]]:
        pass

    def _validate_and_return_sheet_name(self, sheet_name: str) -> str:
        """
        Validates if the given sheet name exists in the workbook and returns it.

        Parameters:
            sheet_name: The name of the sheet to validate.

        Returns:
            The validated sheet name.

        Raises:
            ValueError: If the sheet name is not in the workbook.
        """
        sheetnames_list = self._workbook.sheetnames
        if sheet_name not in sheetnames_list:
            raise ValueError(f'"{sheet_name}" is not in the sheet names. Possible sheet names: {sheetnames_list}.')
        return sheet_name


class OpenConfig(ABC):
    def __init__(self, config_path: str):
        self._config_path = validate_and_return_file_path(config_path, 'json')



